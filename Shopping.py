from selenium import webdriver

from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options

from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

from webdriver_manager.chrome import ChromeDriverManager

import time

import openpyxl

#################################################################################

customService = Service(ChromeDriverManager().install())
customOption = Options()

browser = webdriver.Chrome(service=customService, options=customOption)

URL = 'https://datalab.naver.com/shoppingInsight/sCategory.naver'
browser.get(URL)
browser.implicitly_wait(10)

#################################################################################

class1 = '식품'
class2 = ''
class3 = ''
class4 = ''

# class1
browser.find_element(By.XPATH, '//*[@id="content"]/div[2]/div/div[1]/div/div/div[1]/div/div[1]/span').click()
browser.find_element(By.XPATH,
                     f'//div[2]/div/div[1]/div/div/div[1]/div/div[1]/ul/li/a[contains(text(), "{class1}")]').click()

# class2
if class2 != '':
    browser.find_element(By.XPATH, '//*[@id="content"]/div[2]/div/div[1]/div/div/div[1]/div/div[2]/span').click()
    browser.find_element(By.XPATH,
                         f'//div[2]/div/div[1]/div/div/div[1]/div/div[2]/ul/li/a[contains(text(), "{class2}")]').click()

# class3
if class3 != '':
    browser.find_element(By.XPATH, '//*[@id="content"]/div[2]/div/div[1]/div/div/div[1]/div/div[3]/span').click()
    browser.find_element(By.XPATH,
                         f'//div[2]/div/div[1]/div/div/div[1]/div/div[3]/ul/li/a[contains(text(), "{class3}")]').click()

# class4
if class4 != '':
    browser.find_element(By.XPATH, '//*[@id="content"]/div[2]/div/div[1]/div/div/div[1]/div/div[4]/span').click()
    browser.find_element(By.XPATH,
                         f'//div[2]/div/div[1]/div/div/div[1]/div/div[4]/ul/li/a[contains(text(), "{class4}")]').click()

# 조회하기 클릭
browser.find_element(By.XPATH, '//*[@id="content"]/div[2]/div/div[1]/div/a').click()
time.sleep(3)

#################################################################################

# 엑셀 모듈에서 파일 -> 시트 객체 생성
xlsxFile = openpyxl.Workbook()
xlsxSheet = xlsxFile.active

# TOP500 긁기
for i in range(25):
    for j in range(1, 21, 1):
        keyword = browser.find_element(By.XPATH,
                                       f'//*[@id="content"]/div[2]/div/div[2]/div[2]/div/div/div[1]/ul/li[{j}]/a').text
        print(keyword)
        keyword = keyword.split('\n')

        xlsxSheet.cell(row=(i * 20 + j), column=1).value = (i * 20 + j)
        xlsxSheet.cell(row=(i * 20 + j), column=2).value = keyword[1]

    browser.find_element(By.XPATH, '//*[@id="content"]/div[2]/div/div[2]/div[2]/div/div/div[2]/div/a[2]').click()
    time.sleep(3)

# 엑셀 파일 저장
xlsxFile.save('resultShopping.xlsx')
